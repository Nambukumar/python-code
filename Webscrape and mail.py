import os
import unittest
import smtplib
import requests
import warnings
import time
import numpy as np
import datetime
import re
import xlwings
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from cssselect import GenericTranslator, SelectorError
from bs4 import BeautifulSoup
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.message import EmailMessage

#os.system("taskkill /f /im chrome.exe")

x = datetime.datetime.now()
day = x.day
month = x.strftime("%B")
toshi = x.year
gap = " "
date = str(day) + gap + str(month) + gap + str(toshi)

EMAIL_ADDRESS = os.environ.get('DB_USER')
EMAIL_PASSWORD = os.environ.get('DB_PASS')

recipients = ['sample@testmail.com','sample2@testmail.com']

hcl_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000003I3iwEAC'
cts_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001qFhiEAE'
tcs_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001srGWEAY'
infosys_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001ry7oEAA'
wipro_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001ga8kEAA'
IBM_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001gKHUEA2'
del_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001yKpNEAU'
acc_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N3000000266zBEAQ'
cap_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000003Hub3EAC'
techm_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N30000001rksIEAQ'
livearea_url = 'https://appexchange.salesforce.com/appxConsultingListingDetail?listingId=a0N3A00000EtDunUAF'

###################################################_HCL_################################################################################
browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(hcl_url)
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
hcl_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
hcl_button.click()
time.sleep(8)	
hcl_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
hcl_split_count = hcl_certificate.text
hcl_array_count = hcl_split_count.split(" ")

hcloverall = re.sub("[^0-9]", "" , hcl_array_count[1])
hcl_Admin_Count = re.sub("[^0-9]", "" , hcl_array_count[18])
hcl_Architect_Count = re.sub("[^0-9]", "" , hcl_array_count[20])
hcl_Consultant_Count = re.sub("[^0-9]", "" , hcl_array_count[22])
hcl_Developer_Count = re.sub("[^0-9]", "" , hcl_array_count[24])
hcl_Marketing_Count = re.sub("[^0-9]", "" , hcl_array_count[26])

browser.close()

###################################################_CTS_################################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(cts_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
cts_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
cts_button.click()
time.sleep(8)	
cts_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
cts_split_count = cts_certificate.text
cts_array_count = cts_split_count.split(" ")

ctsoverall = re.sub("[^0-9]", "" , cts_array_count[1])
cts_Admin_Count = re.sub("[^0-9]", "" , cts_array_count[18])
cts_Architect_Count = re.sub("[^0-9]", "" , cts_array_count[20])
cts_Consultant_Count = re.sub("[^0-9]", "" , cts_array_count[22])
cts_Developer_Count = re.sub("[^0-9]", "" , cts_array_count[24])
cts_Marketing_Count = re.sub("[^0-9]", "" , cts_array_count[26])

browser.close()

###################################################_TCS_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(tcs_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
tcs_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
tcs_button.click()
time.sleep(8)	
tcs_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
tcs_split_count = tcs_certificate.text
tcs_array_count = tcs_split_count.split(" ")

tcsoverall = re.sub("[^0-9]", "" , tcs_array_count[1])
tcs_Admin_Count = re.sub("[^0-9]", "" , tcs_array_count[18])
tcs_Architect_Count = re.sub("[^0-9]", "" , tcs_array_count[20])
tcs_Consultant_Count = re.sub("[^0-9]", "" , tcs_array_count[22])
tcs_Developer_Count = re.sub("[^0-9]", "" , tcs_array_count[24])
tcs_Marketing_Count = re.sub("[^0-9]", "" , tcs_array_count[26])

browser.close()

###################################################_infosys_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(infosys_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
infosys_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
infosys_button.click()
time.sleep(8)	
infosys_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
infosys_split_count = infosys_certificate.text
infosys_array_count = infosys_split_count.split(" ")

infosysoverall = re.sub("[^0-9]", "" , infosys_array_count[1])
infosys_Admin_Count = re.sub("[^0-9]", "" , infosys_array_count[18])
infosys_Architect_Count = re.sub("[^0-9]", "" , infosys_array_count[20])
infosys_Consultant_Count = re.sub("[^0-9]", "" , infosys_array_count[22])
infosys_Developer_Count = re.sub("[^0-9]", "" , infosys_array_count[24])
infosys_Marketing_Count = re.sub("[^0-9]", "" , infosys_array_count[26])

browser.close()

###################################################_wipro_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(wipro_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
wipro_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
wipro_button.click()
time.sleep(8)	
wipro_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
wipro_split_count = wipro_certificate.text
wipro_array_count = wipro_split_count.split(" ")

wiprooverall = re.sub("[^0-9]", "" , wipro_array_count[1])
wipro_Admin_Count = re.sub("[^0-9]", "" , wipro_array_count[18])
wipro_Architect_Count = re.sub("[^0-9]", "" , wipro_array_count[20])
wipro_Consultant_Count = re.sub("[^0-9]", "" , wipro_array_count[22])
wipro_Developer_Count = re.sub("[^0-9]", "" , wipro_array_count[24])
wipro_Marketing_Count = re.sub("[^0-9]", "" , wipro_array_count[26])

browser.close()

###################################################_IBM_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(IBM_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
IBM_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
IBM_button.click()
time.sleep(8)	
IBM_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
IBM_split_count = IBM_certificate.text
IBM_array_count = IBM_split_count.split(" ")

IBMoverall = re.sub("[^0-9]", "" , IBM_array_count[1])
IBM_Admin_Count = re.sub("[^0-9]", "" , IBM_array_count[18])
IBM_Architect_Count = re.sub("[^0-9]", "" , IBM_array_count[20])
IBM_Consultant_Count = re.sub("[^0-9]", "" , IBM_array_count[22])
IBM_Developer_Count = re.sub("[^0-9]", "" , IBM_array_count[24])
IBM_Marketing_Count = re.sub("[^0-9]", "" , IBM_array_count[26])

browser.close()

###################################################_DEL_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(del_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
del_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
del_button.click()
time.sleep(8)	
del_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
del_split_count = del_certificate.text
del_array_count = del_split_count.split(" ")

deloverall = re.sub("[^0-9]", "" , del_array_count[1])
del_Admin_Count = re.sub("[^0-9]", "" , del_array_count[18])
del_Architect_Count = re.sub("[^0-9]", "" , del_array_count[20])
del_Consultant_Count = re.sub("[^0-9]", "" , del_array_count[22])
del_Developer_Count = re.sub("[^0-9]", "" , del_array_count[24])
del_Marketing_Count = re.sub("[^0-9]", "" , del_array_count[26])

browser.close()

###################################################_ACC_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(acc_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
acc_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
acc_button.click()
time.sleep(8)	
acc_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
acc_split_count = acc_certificate.text
acc_array_count = acc_split_count.split(" ")

accoverall = re.sub("[^0-9]", "" , acc_array_count[1])
acc_Admin_Count = re.sub("[^0-9]", "" , acc_array_count[18])
acc_Architect_Count = re.sub("[^0-9]", "" , acc_array_count[20])
acc_Consultant_Count = re.sub("[^0-9]", "" , acc_array_count[22])
acc_Developer_Count = re.sub("[^0-9]", "" , acc_array_count[24])
acc_Marketing_Count = re.sub("[^0-9]", "" , acc_array_count[26])

browser.close()

###################################################_CAP_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(cap_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
cap_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
cap_button.click()
time.sleep(8)	
cap_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
cap_split_count = cap_certificate.text
cap_array_count = cap_split_count.split(" ")

capoverall = re.sub("[^0-9]", "" , cap_array_count[1])
cap_Admin_Count = re.sub("[^0-9]", "" , cap_array_count[18])
cap_Architect_Count = re.sub("[^0-9]", "" , cap_array_count[20])
cap_Consultant_Count = re.sub("[^0-9]", "" , cap_array_count[22])
cap_Developer_Count = re.sub("[^0-9]", "" , cap_array_count[24])
cap_Marketing_Count = re.sub("[^0-9]", "" , cap_array_count[26])

browser.close()

###################################################_TECH_M_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(techm_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
techm_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
techm_button.click()
time.sleep(8)	
techm_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
techm_split_count = techm_certificate.text
techm_array_count = techm_split_count.split(" ")

techmoverall = re.sub("[^0-9]", "" , techm_array_count[1])
techm_Admin_Count = re.sub("[^0-9]", "" , techm_array_count[18])
techm_Architect_Count = re.sub("[^0-9]", "" , techm_array_count[20])
techm_Consultant_Count = re.sub("[^0-9]", "" , techm_array_count[22])
techm_Developer_Count = re.sub("[^0-9]", "" , techm_array_count[24])
techm_Marketing_Count = re.sub("[^0-9]", "" , techm_array_count[26])

browser.close()

###################################################_livearea_###############################################################################

browser = webdriver.Chrome(executable_path=ChromeDriverManager().install())
browser.get(livearea_url)	
time.sleep(8)

warnings.filterwarnings("ignore", category=DeprecationWarning) 
livearea_button = browser.find_element_by_css_selector("[aria-controls = tab-default-3]")
livearea_button.click()
time.sleep(8)	
livearea_certificate = browser.find_element_by_css_selector("[id = expersite-certifications-info-id]")
#certificate2 = browser.find_element_by_css_selector("[id = accordion-details-01]")
#print ('Certifications :', certificate.text)
#split_count = np.array(certificate.text)
livearea_split_count = livearea_certificate.text
livearea_array_count = livearea_split_count.split(" ")

liveareaoverall = re.sub("[^0-9]", "" , livearea_array_count[1])
livearea_Admin_Count = re.sub("[^0-9]", "" , livearea_array_count[18])
livearea_Architect_Count = re.sub("[^0-9]", "" , livearea_array_count[20])
livearea_Consultant_Count = re.sub("[^0-9]", "" , livearea_array_count[22])
livearea_Developer_Count = re.sub("[^0-9]", "" , livearea_array_count[24])
livearea_Marketing_Count = re.sub("[^0-9]", "" , livearea_array_count[26])

browser.close()


file_path = "C:\\Users\\nambu\\Desktop\\Live Salesforce Certifications Count.xlsx"
wb = load_workbook(file_path)
ws = wb.worksheets[0]
ws2 = wb.worksheets[1]

######_HCL_Excel_Update###########
ws['B2'].value = hcl_Admin_Count
ws['B3'].value = hcl_Architect_Count
ws['B4'].value = hcl_Consultant_Count
ws['B5'].value = hcl_Developer_Count
ws['B6'].value = hcl_Marketing_Count
ws['B7'].value = hcloverall

######_CTS_Excel_Update###########

ws['C2'].value = cts_Admin_Count
ws['C3'].value = cts_Architect_Count
ws['C4'].value = cts_Consultant_Count
ws['C5'].value = cts_Developer_Count
ws['C6'].value = cts_Marketing_Count
ws['C7'].value = ctsoverall

######_TCS_Excel_Update###########

ws['D2'].value = tcs_Admin_Count
ws['D3'].value = tcs_Architect_Count
ws['D4'].value = tcs_Consultant_Count
ws['D5'].value = tcs_Developer_Count
ws['D6'].value = tcs_Marketing_Count
ws['D7'].value = tcsoverall


######_infosys_Excel_Update###########

ws['E2'].value = infosys_Admin_Count
ws['E3'].value = infosys_Architect_Count
ws['E4'].value = infosys_Consultant_Count
ws['E5'].value = infosys_Developer_Count
ws['E6'].value = infosys_Marketing_Count
ws['E7'].value = infosysoverall


######_wipro_Excel_Update###########

ws['F2'].value = wipro_Admin_Count
ws['F3'].value = wipro_Architect_Count
ws['F4'].value = wipro_Consultant_Count
ws['F5'].value = wipro_Developer_Count
ws['F6'].value = wipro_Marketing_Count
ws['F7'].value = wiprooverall

######_IBM_Excel_Update###########

ws['G2'].value = IBM_Admin_Count
ws['G3'].value = IBM_Architect_Count
ws['G4'].value = IBM_Consultant_Count
ws['G5'].value = IBM_Developer_Count
ws['G6'].value = IBM_Marketing_Count
ws['G7'].value = IBMoverall

######_DEl_Excel_Update###########

ws['H2'].value = del_Admin_Count
ws['H3'].value = del_Architect_Count
ws['H4'].value = del_Consultant_Count
ws['H5'].value = del_Developer_Count
ws['H6'].value = del_Marketing_Count
ws['H7'].value = deloverall

######_DEl_Excel_Update###########

ws['I2'].value = acc_Admin_Count
ws['I3'].value = acc_Architect_Count
ws['I4'].value = acc_Consultant_Count
ws['I5'].value = acc_Developer_Count
ws['I6'].value = acc_Marketing_Count
ws['I7'].value = accoverall

######_CAP_Excel_Update###########

ws['J2'].value = cap_Admin_Count
ws['J3'].value = cap_Architect_Count
ws['J4'].value = cap_Consultant_Count
ws['J5'].value = cap_Developer_Count
ws['J6'].value = cap_Marketing_Count
ws['J7'].value = capoverall

######_TECH_M_Excel_Update###########

ws['K2'].value = techm_Admin_Count
ws['K3'].value = techm_Architect_Count
ws['K4'].value = techm_Consultant_Count
ws['K5'].value = techm_Developer_Count
ws['K6'].value = techm_Marketing_Count
ws['K7'].value = techmoverall

######_livearea_Excel_Update###########

ws['L2'].value = livearea_Admin_Count
ws['L3'].value = livearea_Architect_Count
ws['L4'].value = livearea_Consultant_Count
ws['L5'].value = livearea_Developer_Count
ws['L6'].value = livearea_Marketing_Count
ws['L7'].value = liveareaoverall

wb.save(file_path)

# msg = EmailMessage()
# msg['Subject'] =  str(date) + gap + '[HCL Salesforce CERTIFICATION Count]'
# msg['From'] = EMAIL_ADDRESS
# msg['To'] = recipients
# msg.set_content(hcl_certificate.text)

# with smtplib.SMTP_SSL('smtp.gmail.com',465) as smtp:
# 		smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
# 		smtp.send_message(msg) 


#os.system("taskkill /f /im chrome.exe")	