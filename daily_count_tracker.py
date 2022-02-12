import os
import datetime
import warnings
import time
from openpyxl import Workbook, load_workbook


x = datetime.datetime.now()
day = x.day
month = x.strftime("%B")
toshi = x.year
gap = "-"
date = str(day) + gap + str(month) + gap + str(toshi)

file_path_1 = "C:\\Users\\nambu\\Desktop\\Live Salesforce Certifications Count.xlsx"
wb = load_workbook(file_path_1)
ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]

for a in ws2['A']:
    if(a.value == None):
        da = ws2.cell(row=a.row, column=1, value= date)
        break
    else:
        continue

for b in ws2['B']:
    if(b.value == None):
        db = ws2.cell(row=b.row, column=2, value= ws1['B7'].value)
        break
    else:
        continue

for c in ws2['C']:
    if(c.value == None):
        dc = ws2.cell(row=c.row, column=3, value= ws1['C7'].value)
        break
    else:
        continue

for d in ws2['D']:
    if(d.value == None):
        dd = ws2.cell(row=d.row, column=4, value= ws1['D7'].value)
        break
    else:
        continue

for e in ws2['E']:
    if(e.value == None):
        de = ws2.cell(row=e.row, column=5, value= ws1['E7'].value)
        break
    else:
        continue

for f in ws2['F']:
    if(f.value == None):
        df = ws2.cell(row=f.row, column=6, value= ws1['F7'].value)
        break
    else:
        continue

for g in ws2['G']:
    if(g.value == None):
        dg = ws2.cell(row=g.row, column=7, value= ws1['G7'].value)
        break
    else:
        continue

for h in ws2['H']:
    if(h.value == None):
        dh = ws2.cell(row=h.row, column=8, value= ws1['H7'].value)
        break
    else:
        continue

for i in ws2['I']:
    if(i.value == None):
        di = ws2.cell(row=i.row, column=9, value= ws1['I7'].value)
        break
    else:
        continue

for j in ws2['J']:
    if(j.value == None):
        dj = ws2.cell(row=j.row, column=10, value= ws1['J7'].value)
        break
    else:
        continue

for k in ws2['K']:
    if(k.value == None):
        dk = ws2.cell(row=k.row, column=11, value= ws1['K7'].value)
        break
    else:
        continue

for l in ws2['L']:
    if(l.value == None):
        dl = ws2.cell(row=l.row, column=12, value= ws1['L7'].value)
        break
    else:
        continue

wb.save(file_path_1)